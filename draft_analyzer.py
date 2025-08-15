#!/usr/bin/env python3
"""
Draft Analyzer - Advanced Fantasy Football Analytics
===================================================

Provides deeper analysis including:
- Team situation analysis (coaching changes, O-line, defense)
- Stack analysis and correlation
- Value-based drafting calculations
- Historical performance patterns
- Risk/reward optimization
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple
from dataclasses import dataclass
from fantasy_draft_tool import Player, Position, FantasyDraftTool

@dataclass
class TeamSituation:
    """Team-level situational analysis"""
    team: str
    offensive_coordinator_new: bool = False
    head_coach_new: bool = False
    qb_change: bool = False
    oline_grade_change: float = 0.0  # +/- from previous year
    defense_rank_change: int = 0  # +/- from previous year
    vegas_win_total: float = 8.5
    vegas_point_total: float = 22.5
    pace_rank: int = 16  # 1-32 ranking
    
class SituationalAnalyzer:
    """Analyze situational factors that affect fantasy performance"""
    
    def __init__(self):
        # Historical data based on research
        self.situational_impacts = {
            "new_oc_positive": {
                "QB": 1.15,  # 15% boost average
                "WR": 1.12,  # 12% boost
                "RB": 1.08,  # 8% boost
                "TE": 1.10   # 10% boost
            },
            "improved_oline": {
                "QB": 1.08,
                "RB": 1.18,  # Biggest impact on RBs
                "WR": 1.05,
                "TE": 1.03
            },
            "worse_defense": {
                "QB": 1.12,  # More passing in negative game scripts
                "WR": 1.10,
                "RB": 0.95,  # Less running in negative scripts
                "TE": 1.08
            },
            "contract_year": {
                "QB": 1.05,
                "RB": 1.12,  # Historical biggest impact
                "WR": 1.08,
                "TE": 1.06
            }
        }
        
        # Team situations for 2024 (based on research)
        self.team_situations = {
            "BAL": TeamSituation("BAL", offensive_coordinator_new=True, vegas_win_total=10.5, vegas_point_total=24.5),
            "NYJ": TeamSituation("NYJ", head_coach_new=True, oline_grade_change=8.0, vegas_win_total=9.5),
            "LV": TeamSituation("LV", head_coach_new=True, defense_rank_change=-5, vegas_win_total=7.5),
            "CAR": TeamSituation("CAR", head_coach_new=True, qb_change=True, vegas_win_total=6.5),
            "SEA": TeamSituation("SEA", defense_rank_change=-8, vegas_win_total=8.5),
            "ATL": TeamSituation("ATL", head_coach_new=True, oline_grade_change=6.0, vegas_win_total=8.5),
            "GB": TeamSituation("GB", qb_change=True, oline_grade_change=4.0, vegas_win_total=9.0),
            "SF": TeamSituation("SF", defense_rank_change=3, vegas_win_total=11.5, vegas_point_total=25.0),
            "KC": TeamSituation("KC", vegas_win_total=11.5, vegas_point_total=24.5, pace_rank=8),
            "BUF": TeamSituation("BUF", vegas_win_total=11.0, vegas_point_total=25.5, pace_rank=5)
        }
    
    def analyze_player_situation(self, player: Player) -> Dict[str, float]:
        """Analyze all situational factors for a player"""
        team_situation = self.team_situations.get(player.team)
        if not team_situation:
            return {"base_multiplier": 1.0, "factors": []}
        
        multiplier = 1.0
        factors = []
        
        # New offensive coordinator impact
        if team_situation.offensive_coordinator_new:
            pos_boost = self.situational_impacts["new_oc_positive"].get(player.position.value, 1.0)
            multiplier *= pos_boost
            factors.append(f"New OC: +{(pos_boost-1)*100:.0f}%")
        
        # O-line improvements
        if team_situation.oline_grade_change > 5:
            pos_boost = self.situational_impacts["improved_oline"].get(player.position.value, 1.0)
            multiplier *= pos_boost
            factors.append(f"Improved O-line: +{(pos_boost-1)*100:.0f}%")
        
        # Defensive ranking decline (more passing situations)
        if team_situation.defense_rank_change < -5:
            pos_boost = self.situational_impacts["worse_defense"].get(player.position.value, 1.0)
            multiplier *= pos_boost
            factors.append(f"Worse defense: +{(pos_boost-1)*100:.0f}%")
        
        # Contract year motivation
        if player.contract_year:
            pos_boost = self.situational_impacts["contract_year"].get(player.position.value, 1.0)
            multiplier *= pos_boost
            factors.append(f"Contract year: +{(pos_boost-1)*100:.0f}%")
        
        # Vegas totals impact
        if team_situation.vegas_win_total > 10:
            multiplier *= 1.05
            factors.append("High win total: +5%")
        elif team_situation.vegas_win_total < 7:
            multiplier *= 0.95
            factors.append("Low win total: -5%")
        
        return {
            "base_multiplier": multiplier,
            "factors": factors,
            "vegas_wins": team_situation.vegas_win_total,
            "vegas_points": team_situation.vegas_point_total
        }

class StackAnalyzer:
    """Analyze team stacking opportunities"""
    
    def __init__(self):
        # Historical correlation data
        self.stack_correlations = {
            "QB-WR1": 0.65,  # Strong positive correlation
            "QB-WR2": 0.45,  # Moderate correlation
            "QB-TE": 0.52,   # Good correlation
            "RB-DST": -0.30, # Negative correlation (RB does well, DST struggles)
        }
        
        # High-scoring team combinations
        self.elite_stack_teams = [
            "KC", "BUF", "SF", "MIA", "CIN", "LAR", "GB", "DAL"
        ]
    
    def find_stack_opportunities(self, players: List[Player]) -> List[Dict]:
        """Find optimal stacking opportunities"""
        stack_opportunities = []
        
        # Group players by team
        team_players = {}
        for player in players:
            if player.team not in team_players:
                team_players[player.team] = []
            team_players[player.team].append(player)
        
        # Analyze each team's stack potential
        for team, team_roster in team_players.items():
            if len(team_roster) < 2:
                continue
                
            qbs = [p for p in team_roster if p.position == Position.QB]
            wrs = [p for p in team_roster if p.position == Position.WR]
            tes = [p for p in team_roster if p.position == Position.TE]
            
            # QB-WR stacks
            for qb in qbs:
                for i, wr in enumerate(wrs):
                    correlation = self.stack_correlations["QB-WR1"] if i == 0 else self.stack_correlations["QB-WR2"]
                    
                    stack_value = (qb.draft_value + wr.draft_value) * (1 + correlation * 0.1)
                    
                    stack_opportunities.append({
                        "type": "QB-WR",
                        "team": team,
                        "players": [qb.name, wr.name],
                        "correlation": correlation,
                        "combined_value": stack_value,
                        "elite_team": team in self.elite_stack_teams,
                        "recommendation": self._get_stack_recommendation(qb, wr, correlation)
                    })
            
            # QB-TE stacks
            for qb in qbs:
                for te in tes:
                    correlation = self.stack_correlations["QB-TE"]
                    stack_value = (qb.draft_value + te.draft_value) * (1 + correlation * 0.1)
                    
                    stack_opportunities.append({
                        "type": "QB-TE",
                        "team": team,
                        "players": [qb.name, te.name],
                        "correlation": correlation,
                        "combined_value": stack_value,
                        "elite_team": team in self.elite_stack_teams,
                        "recommendation": self._get_stack_recommendation(qb, te, correlation)
                    })
        
        # Sort by combined value
        stack_opportunities.sort(key=lambda x: x["combined_value"], reverse=True)
        return stack_opportunities[:20]  # Top 20 stacks
    
    def _get_stack_recommendation(self, player1: Player, player2: Player, correlation: float) -> str:
        """Generate stack recommendation"""
        if correlation > 0.6 and player1.championship_frequency > 0.4:
            return "ELITE STACK - High correlation, proven winners"
        elif correlation > 0.5:
            return "STRONG STACK - Good correlation potential"
        elif player1.adp > 50 or player2.adp > 50:
            return "VALUE STACK - Late round opportunity"
        else:
            return "CONSIDER - Moderate upside"

class ValueBasedDrafting:
    """Calculate VBD (Value Based Drafting) scores"""
    
    def __init__(self):
        # Replacement level players by position (12-team league)
        self.replacement_levels = {
            Position.QB: 12,   # 12th QB
            Position.RB: 24,   # 24th RB (2 per team)
            Position.WR: 36,   # 36th WR (3 per team)
            Position.TE: 12    # 12th TE
        }
    
    def calculate_vbd_scores(self, players: List[Player]) -> Dict[str, float]:
        """Calculate VBD scores for all players"""
        vbd_scores = {}
        
        # Group players by position and sort by fantasy points
        position_groups = {}
        for player in players:
            if player.position not in position_groups:
                position_groups[player.position] = []
            position_groups[player.position].append(player)
        
        # Calculate VBD for each position
        for position, pos_players in position_groups.items():
            if position not in self.replacement_levels:
                continue
                
            # Sort by projected fantasy points
            pos_players.sort(key=lambda x: x.fantasy_points, reverse=True)
            
            # Find replacement level
            replacement_index = min(self.replacement_levels[position] - 1, len(pos_players) - 1)
            replacement_points = pos_players[replacement_index].fantasy_points if replacement_index < len(pos_players) else 0
            
            # Calculate VBD for each player
            for player in pos_players:
                vbd_score = player.fantasy_points - replacement_points
                vbd_scores[player.name] = max(0, vbd_score)
        
        return vbd_scores

def create_excel_export(tool: FantasyDraftTool, filename: str = "fantasy_draft_tool.xlsx"):
    """Export comprehensive draft tool to Excel with multiple sheets"""
    
    # Initialize analyzers
    situational = SituationalAnalyzer()
    stack_analyzer = StackAnalyzer()
    vbd_calculator = ValueBasedDrafting()
    
    # Calculate VBD scores
    vbd_scores = vbd_calculator.calculate_vbd_scores(tool.players)
    
    # Main draft sheet
    main_df = tool.generate_draft_sheet()
    
    # Add VBD scores
    main_df['VBD Score'] = main_df['Player'].map(vbd_scores).fillna(0)
    
    # Add situational analysis
    situational_data = []
    for player in tool.players:
        analysis = situational.analyze_player_situation(player)
        situational_data.append({
            'Player': player.name,
            'Position': player.position.value,
            'Team': player.team,
            'Situation Multiplier': f"{analysis['base_multiplier']:.2f}x",
            'Key Factors': " | ".join(analysis['factors']),
            'Vegas Wins': analysis.get('vegas_wins', 'N/A'),
            'Vegas Points': analysis.get('vegas_points', 'N/A'),
            'Adjusted Value': player.draft_value * analysis['base_multiplier']
        })
    
    situational_df = pd.DataFrame(situational_data)
    
    # Stack analysis
    stacks = stack_analyzer.find_stack_opportunities(tool.players)
    stack_df = pd.DataFrame(stacks)
    
    # Round-by-round targets
    round_targets_data = []
    for round_num in range(1, 16):  # 15 rounds
        targets = tool.get_round_targets(round_num)
        for target in targets[:5]:  # Top 5 per round
            round_targets_data.append({
                'Round': round_num,
                'Player': target.name,
                'Position': target.position.value,
                'ADP': target.adp,
                'Draft Value': target.draft_value,
                'Risk Level': target.risk_level.value,
                'Notes': target.notes
            })
    
    round_targets_df = pd.DataFrame(round_targets_data)
    
    # Strategy recommendations
    recommendations = tool.get_strategy_recommendations()
    strategy_df = pd.DataFrame([
        {'Category': k, 'Recommendation': v} 
        for k, v in recommendations.items()
    ])
    
    # Championship team analysis
    championship_data = []
    for player in tool.players:
        if player.championship_frequency > 0.3:  # 30%+ championship frequency
            championship_data.append({
                'Player': player.name,
                'Position': player.position.value,
                'Championship %': f"{player.championship_frequency:.1%}",
                'ADP': player.adp,
                'Draft Value': player.draft_value,
                'Why Successful': player.notes
            })
    
    championship_df = pd.DataFrame(championship_data)
    championship_df = championship_df.sort_values('Championship %', ascending=False)
    
    # Export to Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        main_df.to_excel(writer, sheet_name='Main Draft Board', index=False)
        situational_df.to_excel(writer, sheet_name='Situational Analysis', index=False)
        stack_df.to_excel(writer, sheet_name='Stack Opportunities', index=False)
        round_targets_df.to_excel(writer, sheet_name='Round Targets', index=False)
        strategy_df.to_excel(writer, sheet_name='Strategy Guide', index=False)
        championship_df.to_excel(writer, sheet_name='Championship Players', index=False)
        
        # Format the main draft board
        workbook = writer.book
        worksheet = writer.sheets['Main Draft Board']
        
        # Color code based on tier/risk
        from openpyxl.styles import PatternFill
        
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Apply color coding (simplified for demo)
        for row_num, row in enumerate(main_df.itertuples(), 2):
            if row._9 == "Elite":  # Risk Level column
                for col in range(1, len(main_df.columns) + 1):
                    worksheet.cell(row=row_num, column=col).fill = green_fill
            elif row._9 == "Risky":
                for col in range(1, len(main_df.columns) + 1):
                    worksheet.cell(row=row_num, column=col).fill = red_fill
    
    print(f"✅ Excel export complete: {filename}")
    print(f"📊 Sheets created: Main Draft Board, Situational Analysis, Stack Opportunities, Round Targets, Strategy Guide, Championship Players")
    
    return filename

def main():
    """Demo the draft analyzer"""
    print("📈 Fantasy Draft Analyzer - Advanced Analytics")
    print("=" * 50)
    
    # Initialize main tool
    from fantasy_draft_tool import FantasyDraftTool, LeagueSettings
    
    tool = FantasyDraftTool()
    
    # Configure for a sample league
    league = LeagueSettings(
        name="12-Team PPR",
        scoring="PPR",
        roster_spots={"QB": 1, "RB": 2, "WR": 2, "TE": 1, "FLEX": 1},
        bench_size=6
    )
    
    tool.set_league_settings(league)
    tool.load_player_data({})
    
    # Run situational analysis
    print("\n🔍 Situational Analysis Sample:")
    situational = SituationalAnalyzer()
    
    sample_players = tool.players[:3]
    for player in sample_players:
        analysis = situational.analyze_player_situation(player)
        print(f"\n{player.name} ({player.team}):")
        print(f"  Situation Multiplier: {analysis['base_multiplier']:.2f}x")
        print(f"  Key Factors: {', '.join(analysis['factors']) if analysis['factors'] else 'None'}")
    
    # Stack analysis
    print("\n🏗️ Top Stack Opportunities:")
    stack_analyzer = StackAnalyzer()
    stacks = stack_analyzer.find_stack_opportunities(tool.players)
    
    for i, stack in enumerate(stacks[:5], 1):
        print(f"\n{i}. {stack['team']} {stack['type']}: {' + '.join(stack['players'])}")
        print(f"   Correlation: {stack['correlation']:.2f} | Value: {stack['combined_value']:.1f}")
        print(f"   {stack['recommendation']}")
    
    # Create Excel export
    print(f"\n📊 Creating comprehensive Excel export...")
    filename = create_excel_export(tool)
    
    print(f"\n✅ Draft analyzer complete! Open {filename} to access all sheets.")

if __name__ == "__main__":
    main()